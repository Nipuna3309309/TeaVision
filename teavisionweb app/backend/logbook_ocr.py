import os
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import time, re, warnings, copy
from io import BytesIO
from functools import lru_cache
warnings.filterwarnings('ignore')

import cv2
import numpy as np
from pathlib import Path
from scipy.signal import find_peaks
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from paddleocr import PaddleOCR
import requests
import base64
from dotenv import load_dotenv

# Optional Roboflow Integration
load_dotenv(r"c:\Nipuna\TEST\.env")
ROBOFLOW_API_KEY = os.getenv("ROBOFLOW_API_KEY")

USE_ROBOFLOW_OCR = False  # Set to True to use Roboflow API instead of PaddleOCR

INPUT_FOLDER  = Path("data/images")
OUTPUT_FOLDER = Path("data/output")
TEMPLATE_PATH = Path("data/template/template.xlsx")
OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

SHOW_DEBUG = False

OCR_MIN_CONF      = 0.20
CELL_TARGET_HPX   = 120
CELL_MAX_SCALE    = 8.0
CELL_BORDER_STRIP = 3
CELL_PAD_PX       = 8
NARROW_COLS       = {9, 10, 11, 12, 13}
NARROW_TARGET_HPX = 160
NARROW_MAX_SCALE  = 12.0

BLEED_CONF_THRESHOLD = 0.32
BLEED_PATTERNS = [r"[a-zA-Z0-9],", r"[.\-,!?]{1,3}", r"[.]{1,2}", r"[,]{2,}"]

NUMERIC_COLS      = {0, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}
TEXT_COLS         = {1, 5}
MIN_COLS_EXPECTED = 17
EXPECTED_ROWS_AFTER_SKIP = 65
TEMPLATE_HEADER_ROWS     = 6
TEMPLATE_TOTAL_LABEL     = "TOTAL"
DASH_ONLY = re.compile(r"^[.\-,]+$")

thin  = Side(style="thin",   color="000000")
thick = Side(style="medium", color="000000")
BORDER_THIN   = Border(left=thin, right=thin, top=thin, bottom=thin)
FILL_DATA     = PatternFill("solid", fgColor="FFFDE7")
FILL_TOTAL    = PatternFill("solid", fgColor="BBDEFB")
FILL_LOW_CONF = PatternFill("solid", fgColor="FFF2CC")

COL_NAMES = [
    'No', 'Month', 'Yield_Mo', 'Todate', 'Annual',
    'Mixture', 'N', 'P', 'K', 'Mg/MOP',
    'Zinc', 'Urea', "C'cal", "M'ual", 'Weeding',
    'Rainfall', 'Remarks',
]
COL_NAME_BY_INDEX = {idx: name for idx, name in enumerate(COL_NAMES)}

NUMERIC_TOKEN_PAT = re.compile(r"\d+(?:[.,]\d+)?")
NON_NUMERIC_SPACE_PAT = re.compile(r"[^\d.,\-]")
HEADER_NO_PAT = re.compile(r"^\d{1,2}$")
HEADER_MONTH_PAT = re.compile(
    r"^(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec"
    r"|Bep|Nnv|Auu|Jau|Fcd|Col|Mav|Tun|Sct|Aot|Nec|Fob|Map|Seo"
    r"|PR|PRN)",
    re.IGNORECASE,
)
DATA_NO_PAT = re.compile(r"^\d{1,3}$")
DATA_MONTH_PAT = re.compile(r"^(?:[A-Za-z]{1,4}|PR|PRN|TOTAL)$", re.IGNORECASE)
HEADER_ROW_TOKENS = (
    "annual",
    "application",
    "date",
    "day",
    "dolomite",
    "extent",
    "fertilizer",
    "field",
    "finished",
    "ha",
    "mixture",
    "month",
    "remarks",
    "rain",
    "planting",
    "pruning",
    "seedling",
    "started",
    "style",
    "todate",
    "yield",
    "year",
)

EARLY_EXIT_CONF = 0.80
GOOD_ENOUGH_CONF = 0.60
EMPTY_DARK_RATIO_STRICT = 0.003
EMPTY_DARK_RATIO_SOFT = 0.008
EMPTY_STD_DEV = 11.0

KERNEL_2X2 = np.ones((2, 2), np.uint8)
KERNEL_3X3 = np.ones((3, 3), np.uint8)
SHARPEN_STRONG = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
SHARPEN_LIGHT = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])

def _px_to_col_width(px): return max(1.0, (px - 5) / 7.0)
def _px_to_row_ht(px):    return max(5.0, px * 0.75)

def levenshtein(s1, s2):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m, n = len(s1), len(s2)
    dp = list(range(n+1))
    for i in range(1, m+1):
        prev, dp[0] = dp[0], i
        for j in range(1, n+1):
            prev, dp[j] = dp[j], (prev if s1[i-1]==s2[j-1] else 1 + min(prev, dp[j], dp[j-1]))
    return 1.0 - dp[n] / max(m, n)

OCR_CONFUSION = {
    ('b','s'):0.1,('s','b'):0.1,('0','o'):0.1,('o','0'):0.1,
    ('1','l'):0.2,('l','1'):0.2,('1','i'):0.2,('i','1'):0.2,
    ('u','n'):0.2,('n','u'):0.2,('c','e'):0.2,('e','c'):0.2,
    ('a','u'):0.3,('u','a'):0.3,('v','y'):0.2,('y','v'):0.2,
    ('m','n'):0.2,('n','m'):0.2,('d','a'):0.3,('a','d'):0.3,
    ('f','t'):0.3,('t','f'):0.3,('p','r'):0.2,('r','p'):0.2,
    ('g','q'):0.1,('q','g'):0.1,('6','b'):0.1,('b','6'):0.1,
    ('4','a'):0.3,('a','4'):0.3,
}

def weighted_edit_distance(s1, s2):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m, n = len(s1), len(s2)
    dp = [[0.0]*(n+1) for _ in range(m+1)]
    for i in range(m+1): dp[i][0] = float(i)
    for j in range(n+1): dp[0][j] = float(j)
    for i in range(1, m+1):
        for j in range(1, n+1):
            cost = 0.0 if s1[i-1]==s2[j-1] else OCR_CONFUSION.get((s1[i-1],s2[j-1]),1.0)
            dp[i][j] = min(dp[i-1][j]+1.0, dp[i][j-1]+1.0, dp[i-1][j-1]+cost)
            if i>1 and j>1 and s1[i-1]==s2[j-2] and s1[i-2]==s2[j-1]:
                dp[i][j] = min(dp[i][j], dp[i-2][j-2]+0.5)
    return max(0.0, 1.0 - dp[m][n] / max(m, n))

def jaro_winkler(s1, s2, p=0.1, max_prefix=4):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m1, m2 = len(s1), len(s2)
    match_dist = max(0, max(m1,m2)//2 - 1)
    s1m = [False]*m1; s2m = [False]*m2
    matches = transpositions = 0
    for i in range(m1):
        for j in range(max(0,i-match_dist), min(i+match_dist+1,m2)):
            if s2m[j] or s1[i]!=s2[j]: continue
            s1m[i]=s2m[j]=True; matches+=1; break
    if matches==0: return 0.0
    k=0
    for i in range(m1):
        if not s1m[i]: continue
        while not s2m[k]: k+=1
        if s1[i]!=s2[k]: transpositions+=1
        k+=1
    jaro = (matches/m1 + matches/m2 + (matches-transpositions/2)/matches)/3
    prefix=0
    for i in range(min(max_prefix,m1,m2)):
        if s1[i]==s2[i]: prefix+=1
        else: break
    return jaro + prefix*p*(1-jaro)

def qgram_similarity(s1, s2, q=2):
    s1, s2 = s1.lower(), s2.lower()
    if s1==s2: return 1.0
    if len(s1)<q or len(s2)<q: return float(s1==s2)
    g1=[s1[i:i+q] for i in range(len(s1)-q+1)]
    g2=[s2[i:i+q] for i in range(len(s2)-q+1)]
    c1,c2=Counter(g1),Counter(g2)
    inter=sum((c1&c2).values())
    return 2.0*inter/(len(g1)+len(g2)) if (len(g1)+len(g2))>0 else 0.0

def permterm_similarity(s1, s2):
    s1, s2 = s1.lower(), s2.lower()
    if s1==s2: return 1.0
    c1,c2=Counter(s1),Counter(s2)
    inter=sum((c1&c2).values())
    union=sum((c1|c2).values())
    return inter/union if union>0 else 0.0

FUZZY_WEIGHTS = {'levenshtein':0.25,'weighted_edit':0.35,'jaro_winkler':0.20,'qgram':0.12,'permterm':0.08}

def fuzzy_score(ocr_text, candidate):
    return (FUZZY_WEIGHTS['levenshtein']  *levenshtein(ocr_text,candidate)
           +FUZZY_WEIGHTS['weighted_edit']*weighted_edit_distance(ocr_text,candidate)
           +FUZZY_WEIGHTS['jaro_winkler'] *jaro_winkler(ocr_text,candidate)
           +FUZZY_WEIGHTS['qgram']        *qgram_similarity(ocr_text,candidate)
           +FUZZY_WEIGHTS['permterm']     *permterm_similarity(ocr_text,candidate))

def fuzzy_best_match(ocr_text, vocabulary, min_score=0.40):
    if not ocr_text: return None, 0.0
    t = ocr_text.strip().lower()
    best_cand, best_sc = max(
        ((c, fuzzy_score(t, c.lower())) for c in vocabulary),
        key=lambda x: x[1],
        default=(None, 0.0),
    )
    return (best_cand, best_sc) if best_sc>=min_score else (None, 0.0)

MONTH_VOCAB = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','PR','PRN']
MONTH_VOCAB_EXTENDED = MONTH_VOCAB + ['January','February','March','April','May','June',
                                       'July','August','September','October','November','December']
MONTH_CANONICAL = {
    'January':'Jan','February':'Feb','March':'Mar','April':'Apr',
    'May':'May','June':'Jun','July':'Jul','August':'Aug',
    'September':'Sep','October':'Oct','November':'Nov','December':'Dec',
    'Jan':'Jan','Feb':'Feb','Mar':'Mar','Apr':'Apr',
    'Jun':'Jun','Jul':'Jul','Aug':'Aug','Sep':'Sep',
    'Oct':'Oct','Nov':'Nov','Dec':'Dec','PR':'PR','PRN':'PRN',
}

FERT_VOCAB = ['U-709','U-330','U-428','T-750','B/A','B/N','DAP','MOP',
              'PH-B','PH-4.99A','PH-5.24A','UT-1125']
FERT_NUM_PAT    = re.compile(r"([Uu][_\-]?|[Tt][_\-]?)(\d{2,4})")
FERT_KNOWN_NUMS = {'709':'U-709','330':'U-330','428':'U-428','750':'T-750','1125':'UT-1125'}

@lru_cache(maxsize=256)
def _fuzzy_match_month_cached(text):
    t = str(text).strip()
    if not t: return t
    tl = t.lower()
    for m in MONTH_VOCAB_EXTENDED:
        if tl==m.lower(): return MONTH_CANONICAL.get(m, m[:3].title())
    if len(tl)>=3:
        for m in MONTH_VOCAB_EXTENDED:
            if m.lower().startswith(tl[:3]) or tl[:3]==m.lower()[:3]:
                return MONTH_CANONICAL.get(m, m[:3].title())
    best,sc = fuzzy_best_match(tl, MONTH_VOCAB_EXTENDED, min_score=0.42)
    if best: return MONTH_CANONICAL.get(best, best[:3].title())
    return t[:3].title() if t[:3].isalpha() else t

def fuzzy_match_month(ocr_text):
    return _fuzzy_match_month_cached(str(ocr_text).strip())

@lru_cache(maxsize=512)
def _fuzzy_match_fert_cached(text):
    t = str(text).strip()
    if not t: return t
    m = FERT_NUM_PAT.search(t)
    if m:
        num=m.group(2); prefix=m.group(1).upper().rstrip('-_')
        if num in FERT_KNOWN_NUMS: return FERT_KNOWN_NUMS[num]
        return f'{prefix}-{num}'
    digits = re.findall(r"\d{3,4}", t)
    for d in digits:
        if d in FERT_KNOWN_NUMS: return FERT_KNOWN_NUMS[d]
    best,sc = fuzzy_best_match(t, FERT_VOCAB, min_score=0.38)
    if best: return best
    tl = t.lower()
    if re.search(r"b[/\\]?a", tl): return "B/A"
    if re.search(r"b[/\\]?n", tl): return "B/N"
    if "dap" in tl: return "DAP"
    if "mop" in tl: return "MOP"
    if re.search(r"ph[-\s]?b", tl): return "PH-B"
    return ' '.join(t.split())

def fuzzy_match_fert(ocr_text):
    return _fuzzy_match_fert_cached(str(ocr_text).strip())

REC_MODEL_DIR = r"C:\rec"

OCR_ENGINE = None

def make_ocr_engine():
    if USE_ROBOFLOW_OCR:
        if not ROBOFLOW_API_KEY:
            raise RuntimeError("ROBOFLOW_API_KEY is not defined in .env")
        print('Roboflow OCR endpoint enabled.')
        return "ROBOFLOW"
    
    for kwargs in [
        dict(lang="en", det_model_name="PP-OCRv5_server_det", rec_model_dir=REC_MODEL_DIR, use_textline_orientation=False),
        dict(lang="en", ocr_version="PP-OCRv5", use_textline_orientation=False),
        dict(lang="en", use_textline_orientation=False),
        dict(lang="en"),
    ]:
        try:
            eng = PaddleOCR(**kwargs)
            print(f'PaddleOCR loaded {list(kwargs.keys())}')
            return eng
        except Exception:
            continue
    raise RuntimeError("PaddleOCR init failed")

def init_ocr_engine():
    global OCR_ENGINE
    if OCR_ENGINE is None:
        OCR_ENGINE = make_ocr_engine()
        print('OCR engine ready.')
    return OCR_ENGINE

def parse_paddle_output(out):
    items = []
    if not out: return items
    if isinstance(out, list) and len(out)>0 and isinstance(out[0], dict):
        d = out[0]
        texts  = d.get("rec_texts") or d.get("texts") or []
        scores = d.get("rec_scores") or d.get("scores") or [1.0]*len(texts)
        boxes  = d.get("dt_polys") or d.get("rec_polys") or d.get("dt_boxes") or d.get("boxes") or []
        if isinstance(boxes, np.ndarray): boxes = boxes.tolist()
        if isinstance(texts, str): texts = [texts]
        for b,t,s in zip(boxes, texts, scores): items.append((b, t, float(s)))
        return items
    lines = out[0] if isinstance(out[0], list) and not isinstance(out[0][0], list) else out
    for line in lines:
        try: items.append((line[0], line[1][0], float(line[1][1])))
        except Exception: pass
    return items

def run_ocr(img_bgr):
    if img_bgr is None or img_bgr.size==0: return []
    engine = init_ocr_engine()
    
    if engine == "ROBOFLOW":
        if not ROBOFLOW_API_KEY: return []
        success, buffer = cv2.imencode('.jpg', img_bgr)
        if not success: return []
        img_str = base64.b64encode(buffer).decode("utf-8")
        
        url = f"https://infer.roboflow.com/doctr/ocr?api_key={ROBOFLOW_API_KEY}"
        try:
            resp = requests.post(url, json={"image": {"type": "base64", "value": img_str}}, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                res_text = data.get("result", "")
                if res_text:
                    dummy_box = [[0,0], [img_bgr.shape[1],0], [img_bgr.shape[1],img_bgr.shape[0]], [0,img_bgr.shape[0]]]
                    return [(dummy_box, res_text.strip(), 1.0)]
        except Exception as e:
            print(f" OCR API Error: {e}")
        return []
        
    else:
        if img_bgr.ndim==2: img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)
        try:
            if hasattr(engine, 'predict'):
                try:
                    out = engine.predict(img_bgr, use_doc_orientation_classify=False,
                                         use_doc_unwarping=False, use_textline_orientation=False)
                except TypeError:
                    out = engine.predict(img_bgr)
            else:
                out = engine.ocr(img_bgr)
            return parse_paddle_output(out)
        except Exception:
            return []

def four_point_warp(bgr, pts):
    pts=pts.reshape(4,2); s=pts.sum(1); d=np.diff(pts,axis=1)
    rect=np.array([pts[s.argmin()],pts[d.argmin()],pts[s.argmax()],pts[d.argmax()]],dtype=np.float32)
    tl,tr,br,bl=rect
    W=int(max(np.linalg.norm(br-bl),np.linalg.norm(tr-tl)))
    H=int(max(np.linalg.norm(tr-br),np.linalg.norm(tl-bl)))
    dst=np.array([[0,0],[W-1,0],[W-1,H-1],[0,H-1]],dtype=np.float32)
    return cv2.warpPerspective(bgr,cv2.getPerspectiveTransform(rect,dst),(W,H))

def rectify_document(bgr):
    gray=cv2.cvtColor(bgr,cv2.COLOR_BGR2GRAY)
    edges=cv2.Canny(cv2.GaussianBlur(gray,(5,5),0),50,150)
    edges=cv2.dilate(edges,np.ones((5,5),np.uint8))
    cnts,_=cv2.findContours(edges,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    if not cnts: return bgr
    for c in sorted(cnts,key=cv2.contourArea,reverse=True)[:5]:
        ap=cv2.approxPolyDP(c,0.02*cv2.arcLength(c,True),True)
        if len(ap)==4: return four_point_warp(bgr,ap)
    return bgr

def split_left_right(bgr, pad=3):
    h, w = bgr.shape[:2]
    _, bi = cv2.threshold(cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY),
                          200, 255, cv2.THRESH_BINARY_INV)
    vk = max(30, h//18)
    vm = cv2.morphologyEx(bi, cv2.MORPH_OPEN,
                          cv2.getStructuringElement(cv2.MORPH_RECT, (1, vk)))
    proj = np.convolve(
        np.sum(vm[int(h*0.2):, :] > 0, axis=0).astype(float),
        np.ones(max(15, w//200)) / max(15, w//200),
        mode='same',
    )
    if proj.max() == 0:
        return [(1, bgr)]
    peaks, _ = find_peaks(proj, height=proj.max()*0.25,
                          distance=max(12, w//100))
    if len(peaks) == 0:
        return [(1, bgr)]
    mid = peaks[np.argmin(np.abs(peaks - w*0.5))]
    tables = [(i, bgr[:, max(0, x1-pad):min(w, x2+pad)].copy())
              for i, (x1, x2) in enumerate([(0, mid), (mid, w)], 1)
              if x2 - x1 > w*0.1]
    return tables if len(tables) == 2 else [(1, bgr)]

def cluster_pos(pos, eps=8):
    pos=np.sort(np.unique(pos.astype(int)))
    if len(pos)==0: return pos
    groups=[[pos[0]]]
    for p in pos[1:]:
        if p-groups[-1][-1]<=eps: groups[-1].append(p)
        else: groups.append([p])
    return np.array([int(np.mean(g)) for g in groups])

def detect_grid(bgr):
    h,w=bgr.shape[:2]
    g=cv2.createCLAHE(2.0,(8,8)).apply(cv2.cvtColor(bgr,cv2.COLOR_BGR2GRAY))
    th=cv2.adaptiveThreshold(g,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY_INV,31,15)
    th=cv2.morphologyEx(th,cv2.MORPH_CLOSE,np.ones((3,3),np.uint8),iterations=2)
    hk=max(20,w//30); vk=max(20,h//30)
    hl=cv2.dilate(cv2.morphologyEx(th,cv2.MORPH_OPEN,cv2.getStructuringElement(cv2.MORPH_RECT,(hk,1))),np.ones((1,3),np.uint8))
    vl=cv2.dilate(cv2.morphologyEx(th,cv2.MORPH_OPEN,cv2.getStructuringElement(cv2.MORPH_RECT,(1,vk))),np.ones((3,1),np.uint8))
    ys=[y+hh//2 for c in cv2.findContours(hl,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)[0]
        for x,y,ww,hh in [cv2.boundingRect(c)] if ww>=w*0.40]
    xs=[x+ww//2 for c in cv2.findContours(vl,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)[0]
        for x,y,ww,hh in [cv2.boundingRect(c)] if hh>=h*0.40]
    ys=cluster_pos(np.array(ys)); xs=cluster_pos(np.array(xs))
    def border(lines,lo,hi,tol=14):
        l=list(lines)
        if not l or abs(l[0]-lo)>tol: l=[lo]+l
        if abs(l[-1]-hi)>tol: l=l+[hi]
        return np.array(sorted(l))
    return border(ys,0,h-1), border(xs,0,w-1)

def enforce_min_columns(vlines, min_cols=MIN_COLS_EXPECTED):
    vl=list(vlines)
    while len(vl)-1<min_cols:
        widths=[(vl[i+1]-vl[i],i) for i in range(len(vl)-1)]
        idx=max(widths)[1]
        mid=(vl[idx]+vl[idx+1])//2
        vl.insert(idx+1,mid)
    return np.array(vl,dtype=float)

def split_three_tables(bgr, pad=3):
    h,w=bgr.shape[:2]
    bi=cv2.threshold(cv2.cvtColor(bgr,cv2.COLOR_BGR2GRAY),200,255,cv2.THRESH_BINARY_INV)[1]
    vk=max(int(h/20),30)
    vertclean=cv2.morphologyEx(bi,cv2.MORPH_OPEN,cv2.getStructuringElement(cv2.MORPH_RECT,(1,vk)))
    proj=np.sum(vertclean,axis=0).astype(float)
    if proj.max()==0: return [(1,bgr)]
    peaks,_=find_peaks(proj,height=proj.max()*0.10,distance=5)
    zones={'A':(int(w*0.25),int(w*0.40)),'B':(int(w*0.45),int(w*0.55)),'C':(int(w*0.60),int(w*0.75))}
    noise_floor=w*0.015; best={}
    for zn,(zx0,zx1) in zones.items():
        zp=[p for p in peaks if zx0<p<zx1]
        if len(zp)<2: best[zn]={'size':0,'cutx':0}; continue
        gaps=np.diff(zp); idx=int(np.argmax(gaps))
        best[zn]={'size':int(gaps[idx]),'cutx':int(zp[idx+1])}
    cuts=[]
    if best['A']['size']>noise_floor and best['A']['size']>best['B']['size']:
        cuts.append(best['A']['cutx'])
        if best['C']['size']>noise_floor: cuts.append(best['C']['cutx'])
        else: cuts.append(int(w*0.66))
    else:
        cx=best['B']['cutx'] if best['B']['size']>noise_floor else int(w*0.50)
        cuts.append(cx)
    boundaries=[0]+sorted(cuts)+[w]
    tables=[(i+1,bgr[:,max(0,x1-pad):min(w,x2+pad)].copy())
            for i,(x1,x2) in enumerate(zip(boundaries,boundaries[1:])) if x2-x1>w*0.10]
    return tables if tables else [(1,bgr)]

def auto_split_tables(bgr, pad=3):
    h,w=bgr.shape[:2]
    bi=cv2.threshold(cv2.cvtColor(bgr,cv2.COLOR_BGR2GRAY),200,255,cv2.THRESH_BINARY_INV)[1]
    vk=max(int(h/20),30)
    vertclean=cv2.morphologyEx(bi,cv2.MORPH_OPEN,cv2.getStructuringElement(cv2.MORPH_RECT,(1,vk)))
    proj=np.sum(vertclean,axis=0).astype(float)
    if proj.max()==0: return '2-table',[(1,bgr)]
    peaks,_=find_peaks(proj,height=proj.max()*0.10,distance=5)
    noise_floor=w*0.04
    def zone_gap(zx0,zx1):
        zp=sorted([p for p in peaks if zx0<p<zx1])
        if len(zp)<2: return 0
        return int(np.max(np.diff(zp)))
    gapA=zone_gap(int(w*0.25),int(w*0.40))
    gapB=zone_gap(int(w*0.45),int(w*0.55))
    if gapA>noise_floor and gapA>gapB*1.5:
        layout='3-table'; tables=split_three_tables(bgr,pad=pad)
    else:
        layout='2-table'; tables=split_left_right(bgr,pad=pad)
    print(f'  Layout detected: {layout}  (gap_A={gapA} gap_B={gapB} noise_floor={int(noise_floor)})')
    return layout, tables

def is_bleed_through(text, conf, colidx):
    t=str(text).strip()
    if colidx in NUMERIC_COLS and DASH_ONLY.match(t): return True
    if conf<BLEED_CONF_THRESHOLD:
        for pattern in BLEED_PATTERNS:
            if re.search(pattern,t): return True
        if colidx==1 and (len(t)>8 or (len(t)<=2 and conf<0.4)): return True
        if colidx in NUMERIC_COLS and not any(c.isdigit() for c in t) and conf<0.4: return True
    return False

def scale_up(gray, target_h=CELL_TARGET_HPX, max_scale=CELL_MAX_SCALE):
    h=gray.shape[0]
    if h<4: return gray
    scale=min(max_scale,max(1.0,target_h/h))
    if scale<1.01: return gray
    return cv2.resize(gray,None,fx=scale,fy=scale,interpolation=cv2.INTER_CUBIC)

def scale_up_narrow(gray):
    h=gray.shape[0]
    if h<4: return gray
    scale=min(NARROW_MAX_SCALE,max(1.0,NARROW_TARGET_HPX/h))
    if scale<1.01: return gray
    return cv2.resize(gray,None,fx=scale,fy=scale,interpolation=cv2.INTER_LANCZOS4)

def prep_numeric(gray):
    g=cv2.createCLAHE(2.0,(4,4)).apply(gray)
    _,b=cv2.threshold(g,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    if np.mean(b)>127: b=cv2.bitwise_not(b)
    b=cv2.fastNlMeansDenoising(b,h=10)
    b=cv2.morphologyEx(b,cv2.MORPH_CLOSE,KERNEL_2X2)
    return cv2.filter2D(b,-1,SHARPEN_STRONG)

def prep_text(gray):
    g=cv2.createCLAHE(3.0,(4,4)).apply(gray)
    b=cv2.adaptiveThreshold(g,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,21,10)
    if np.mean(b)>127: b=cv2.bitwise_not(b)
    return cv2.morphologyEx(b,cv2.MORPH_CLOSE,KERNEL_2X2)

def prep_clahe_only(gray):
    g=cv2.createCLAHE(2.5,(4,4)).apply(gray)
    return cv2.filter2D(g,-1,SHARPEN_LIGHT)

def prep_narrow_col(gray):
    g=cv2.createCLAHE(3.0,(4,4)).apply(gray)
    g=cv2.bilateralFilter(g,5,75,75)
    _,b=cv2.threshold(g,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    if np.mean(b)>127: b=cv2.bitwise_not(b)
    b=cv2.morphologyEx(b,cv2.MORPH_CLOSE,KERNEL_2X2)
    return cv2.filter2D(b,-1,SHARPEN_STRONG)

def build_otsu_variant(gray):
    cg=cv2.createCLAHE(2.0,(4,4)).apply(gray)
    _,otsu=cv2.threshold(cg,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    if np.mean(otsu)>127: otsu=cv2.bitwise_not(otsu)
    return otsu

def pad_gray_crop(gray_crop):
    if gray_crop is None or gray_crop.size==0: return None
    p=CELL_PAD_PX
    return cv2.copyMakeBorder(gray_crop,p,p,p,p,cv2.BORDER_CONSTANT,value=255)

def cell_has_visible_content(gray_crop):
    if gray_crop is None or gray_crop.size==0: return False
    h,w=gray_crop.shape[:2]
    if h<4 or w<4: return False
    ypad=min(max(1,h//10),4)
    xpad=min(max(1,w//12),4)
    if h>2*ypad and w>2*xpad:
        inner=gray_crop[ypad:h-ypad,xpad:w-xpad]
    else:
        inner=gray_crop
    if inner.size==0: return False
    dark_ratio=float(np.count_nonzero(inner<245))/float(inner.size)
    if dark_ratio>=EMPTY_DARK_RATIO_SOFT: return True
    return dark_ratio>=EMPTY_DARK_RATIO_STRICT and float(inner.std())>=EMPTY_STD_DEV

def text_looks_plausible(text, colidx):
    t=str(text).strip()
    if not t: return False
    if colidx in NUMERIC_COLS: return any(ch.isdigit() for ch in t)
    if colidx==1: return any(ch.isalpha() for ch in t)
    if colidx==5: return any(ch.isalnum() for ch in t)
    return len(t)>=2 or any(ch.isalnum() for ch in t)

def ocr_single_cell(gray_crop, colidx):
    if gray_crop is None or gray_crop.size==0: return '',0.0
    if colidx in NARROW_COLS: gray=scale_up_narrow(gray_crop)
    else:                     gray=scale_up(gray_crop)
    otsu_variant=build_otsu_variant(gray)
    if colidx in NARROW_COLS:
        variants=[prep_narrow_col(gray),prep_numeric(gray),prep_clahe_only(gray)]
    elif colidx in NUMERIC_COLS:
        variants=[prep_numeric(gray),prep_clahe_only(gray),otsu_variant]
    else:
        variants=[prep_text(gray),prep_clahe_only(gray),otsu_variant]
    best_text,best_conf='',0.0
    for idx,v in enumerate(variants):
        img=cv2.cvtColor(v,cv2.COLOR_GRAY2BGR) if v.ndim==2 else v
        items=run_ocr(img)
        good=[(t,s) for _,t,s in items if s>=OCR_MIN_CONF]
        if not good: continue
        text=' '.join(t for t,_ in good).strip()
        conf=float(np.mean([s for _,s in good]))
        if conf>best_conf: best_conf,best_text=conf,text
        if conf>=EARLY_EXIT_CONF and text_looks_plausible(text,colidx): break
        if idx==0 and conf>=GOOD_ENOUGH_CONF and text_looks_plausible(text,colidx): break
    if best_text and is_bleed_through(best_text,best_conf,colidx): return '',0.0
    return best_text,best_conf

def extract_all_cells(bgr, hlines, vlines):
    nrows=len(hlines)-1; ncols=len(vlines)-1
    print(f'  Cell OCR: {nrows}x{ncols} = {nrows*ncols} cells')
    gray=cv2.cvtColor(bgr,cv2.COLOR_BGR2GRAY)
    H,W=gray.shape[:2]
    row_bounds=[(min(H-1,int(hlines[r])+CELL_BORDER_STRIP), max(0,int(hlines[r+1])-CELL_BORDER_STRIP))
                for r in range(nrows)]
    col_bounds=[(min(W-1,int(vlines[c])+CELL_BORDER_STRIP), max(0,int(vlines[c+1])-CELL_BORDER_STRIP))
                for c in range(ncols)]
    results={}; filtered_bleed=0; skipped_empty=0; t0=time.time()
    for r,(y1,y2) in enumerate(row_bounds):
        if y2-y1<4:
            skipped_empty+=ncols
            continue
        for c,(x1,x2) in enumerate(col_bounds):
            if x2-x1<4:
                skipped_empty+=1
                continue
            crop=gray[y1:y2,x1:x2]
            if not cell_has_visible_content(crop):
                skipped_empty+=1
                continue
            text,conf=ocr_single_cell(pad_gray_crop(crop),c)
            if text:
                results[(r,c)]={'text':text,'conf':conf}
            elif conf==0.0:
                filtered_bleed+=1
        if (r+1)%10==0 or r==nrows-1:
            print(f'  Row {r+1}/{nrows} | {time.time()-t0:.0f}s')
    print(f'  Done {len(results)}/{nrows*ncols} empty={skipped_empty} bleed={filtered_bleed} {time.time()-t0:.1f}s')
    return results, nrows, ncols

def clean_numeric_text(text):
    text=str(text).strip()
    if not text or DASH_ONLY.match(text): return ''
    spaced=NON_NUMERIC_SPACE_PAT.sub(" ",text)
    nums=NUMERIC_TOKEN_PAT.findall(spaced)
    if nums: return nums[0] if len(nums)==1 else '.'.join(nums)
    nums2=NUMERIC_TOKEN_PAT.findall(text)
    if nums2: return nums2[0] if len(nums2)==1 else '.'.join(nums2)
    return ''.join(c for c in text if c.isdigit() or c in '.,-.').strip()

def clean_cell_text(colidx, text):
    text=str(text).strip()
    if not text: return ''
    if colidx in NUMERIC_COLS and DASH_ONLY.match(text): return ''
    if colidx in NUMERIC_COLS: return clean_numeric_text(text)
    if colidx==1: return fuzzy_match_month(text)
    if colidx==5: return fuzzy_match_fert(text)
    return ' '.join(text.split())

def postprocess_grid(raw_results, nrows, ncols):
    cleaned={}
    for (r,c),info in raw_results.items():
        val=clean_cell_text(c,info['text'])
        if val: cleaned[(r,c)]=val
    return cleaned

def detect_data_start(raw_results, nrows, max_scan=15, fallback=5):
    skip_grid=max(0,nrows-EXPECTED_ROWS_AFTER_SKIP)
    skip_ocr=None
    for r in range(min(nrows,max_scan)):
        t0=str(raw_results.get((r,0),{}).get('text','')).strip()
        t1=str(raw_results.get((r,1),{}).get('text','')).strip()
        has_no=bool(HEADER_NO_PAT.match(t0)); has_month=bool(HEADER_MONTH_PAT.match(t1))
        n_filled=sum(1 for c in range(len(COL_NAMES)) if raw_results.get((r,c),{}).get('text','').strip())
        triggered=(has_no and has_month) or (has_no and n_filled>=3) or (has_month and n_filled>=4)
        if triggered:
            row_num=int(t0) if has_no and t0.isdigit() else 1
            skip_ocr=max(0,r-(row_num-1)); break
    if skip_ocr is None:
        skip_ocr=fallback; ocr_label=f'fallback={fallback}'
    else:
        ocr_label=str(skip_ocr)
    if skip_grid==0:
        skip=skip_ocr; reason='grid=0, using OCR'
    elif abs(skip_grid-skip_ocr)<=1:
        skip=skip_grid; reason='methods agree'
    else:
        skip=skip_ocr; reason=f'mismatch(grid={skip_grid},ocr={ocr_label}), using OCR'
    print(f'  DynSkip nrows={nrows} grid={skip_grid} ocr={ocr_label} -> skip={skip} {reason}')
    return skip

def apply_scan_header_skip(cells, conf_dict, skip):
    new_cells={}; new_conf={}
    for (r,c),val in cells.items():
        if r<skip: continue
        r2=r-skip; new_cells[(r2,c)]=val
        if (r,c) in conf_dict: new_conf[(r2,c)]=conf_dict[(r,c)]
    max_r=max((r for r,c in new_cells),default=-1)+1
    print(f'  Skipped {skip} header rows | {max_r} data rows | {len(new_cells)} cells')
    return new_cells, new_conf

def shift_cells_up(cells, conf_dict, skip):
    if skip<=0:
        return cells, conf_dict
    new_cells={}; new_conf={}
    for (r,c),val in cells.items():
        if r<skip:
            continue
        key=(r-skip,c)
        new_cells[key]=val
        if (r,c) in conf_dict:
            new_conf[key]=conf_dict[(r,c)]
    return new_cells, new_conf

def row_values(cells, row_idx, ncols):
    return [str(cells.get((row_idx,c),'')).strip() for c in range(ncols)]

def row_text(vals):
    return ' '.join(v.lower() for v in vals if v)

def is_explicit_header_row(vals):
    joined=row_text(vals)
    if not joined:
        return False
    token_hits=sum(1 for token in HEADER_ROW_TOKENS if token in joined)
    if token_hits>=2:
        return True
    return (
        ('no' in joined and 'month' in joined) or
        ('yield' in joined and 'mixture' in joined) or
        ('fertil' in joined and ('n' in joined or 'p' in joined or 'k' in joined))
    )

def is_data_like_row(vals):
    no_v=vals[0] if vals else ''
    month_v=vals[1] if len(vals)>1 else ''
    has_no=bool(DATA_NO_PAT.fullmatch(no_v))
    has_month=bool(DATA_MONTH_PAT.fullmatch(month_v))
    n_filled=sum(1 for v in vals if v)
    is_total=any(TEMPLATE_TOTAL_LABEL in v.upper() for v in vals if v)
    return (has_no and has_month) or (has_no and n_filled>=3) or (has_month and n_filled>=4) or is_total

def trim_leading_non_data_rows(cells, conf_dict, ncols, max_scan=12):
    if not cells:
        return cells, conf_dict, 0
    max_r=max(r for r,c in cells)
    scan_rows=min(max_r+1, max_scan)
    explicit_skip=0
    while explicit_skip<scan_rows:
        vals=row_values(cells, explicit_skip, ncols)
        if not is_explicit_header_row(vals):
            break
        explicit_skip+=1
    if explicit_skip>0:
        cells, conf_dict = shift_cells_up(cells, conf_dict, explicit_skip)
        max_r=max((r for r,c in cells), default=-1)
        scan_rows=min(max_r+1, max_scan)
    stable_skip=0
    for r in range(scan_rows):
        vals=row_values(cells, r, ncols)
        if is_explicit_header_row(vals):
            stable_skip=r+1
            continue
        if not any(vals):
            stable_skip=r+1
            continue
        row_is_data=is_data_like_row(vals)
        future_hits=0
        for rr in range(r, min(scan_rows, r+5)):
            if is_data_like_row(row_values(cells, rr, ncols)):
                future_hits+=1
        if row_is_data and future_hits>=2:
            stable_skip=r
            break
        if future_hits>=2:
            stable_skip=r+1
    total_skip=explicit_skip+stable_skip
    if stable_skip>0:
        cells, conf_dict = shift_cells_up(cells, conf_dict, stable_skip)
    if total_skip>0:
        print(f'  Extra leading-row trim: {total_skip} rows')
    return cells, conf_dict, total_skip

def find_data_extent(cells, ncols, max_empty_streak=8):
    if not cells: return 0
    max_r=max(r for r,c in cells)
    last_valid=-1; streak=0
    for r in range(max_r+1):
        month_v=str(cells.get((r,1),'')).strip()
        no_v   =str(cells.get((r,0),'')).strip()
        has_no   =bool(DATA_NO_PAT.fullmatch(no_v))
        has_month=bool(DATA_MONTH_PAT.fullmatch(month_v))
        is_total ='TOTAL' in month_v.upper() or 'TOTAL' in no_v.upper()
        n_cells=sum(1 for c in range(ncols) if cells.get((r,c),''))
        if has_no or has_month or (is_total and n_cells>=1):
            last_valid=r; streak=0
        else:
            streak+=1
            if streak>=max_empty_streak and last_valid>=0: break
    return max(last_valid,0)

def trim_data_region(cells, conf_dict, hlines, vlines, min_row_fill=0.05, min_col_fill=0.05):
    nrows=len(hlines)-1; ncols=len(vlines)-1
    if nrows==0 or ncols==0: return cells,conf_dict,hlines,vlines
    row_cnt=[sum(1 for c in range(ncols) if cells.get((r,c),'')) for r in range(nrows)]
    col_cnt=[sum(1 for r in range(nrows) if cells.get((r,c),'')) for c in range(ncols)]
    r_thresh=max(1,int(ncols*min_row_fill)); c_thresh=max(1,int(nrows*min_col_fill))
    act_r=[i for i,v in enumerate(row_cnt) if v>=r_thresh]
    act_c=[i for i,v in enumerate(col_cnt) if v>=c_thresh]
    if not act_r or not act_c: return cells,conf_dict,hlines,vlines
    r0,r1=min(act_r),max(act_r); c0,c1=min(act_c),max(act_c)
    new_cells={}; new_conf={}
    for (r,c),val in cells.items():
        if r0<=r<=r1 and c0<=c<=c1:
            key=(r-r0,c-c0)
            new_cells[key]=val
            if (r,c) in conf_dict: new_conf[key]=conf_dict[(r,c)]
    return new_cells, new_conf, hlines[r0:r1+2], vlines[c0:c1+2]

def write_filled_template(cells_by_table, conf_by_table, col_widths_by_table, row_hts_by_table, template_path, out_path):
    ntables=len(cells_by_table)
    side_labels={1:'LeftTable',2:'RightTable'} if ntables==2 else {i:f'Table{i}' for i in cells_by_table}
    try:
        template_bytes=Path(template_path).read_bytes()
    except Exception:
        template_bytes=None
    out_wb=Workbook(); out_wb.remove(out_wb.active)
    for tidx,cells in cells_by_table.items():
        side=side_labels.get(tidx,f'Table{tidx}')
        try:
            if template_bytes is not None:
                tmpl_wb=load_workbook(BytesIO(template_bytes))
            else:
                raise FileNotFoundError
            ws_src=tmpl_wb.active
        except Exception:
            tmpl_wb=Workbook(); ws_src=tmpl_wb.active
        if tidx in col_widths_by_table:
            for ci,w in enumerate(col_widths_by_table[tidx]):
                ws_src.column_dimensions[get_column_letter(ci+1)].width=_px_to_col_width(w)
        if tidx in row_hts_by_table:
            for ri,h in enumerate(row_hts_by_table[tidx]):
                ws_src.row_dimensions[TEMPLATE_HEADER_ROWS+ri+1].height=_px_to_row_ht(h)
        cd=conf_by_table.get(tidx,{})
        for (r,c),val in cells.items():
            excel_row=TEMPLATE_HEADER_ROWS+r+1
            if excel_row<=TEMPLATE_HEADER_ROWS: continue
            cell=ws_src.cell(row=excel_row,column=c+1,value=val)
            cell.border=BORDER_THIN
            is_total=str(val).upper().strip()==TEMPLATE_TOTAL_LABEL
            if is_total:
                cell.fill=FILL_TOTAL; cell.font=Font(bold=True)
                cell.alignment=Alignment(horizontal='center',vertical='center')
            else:
                conf_val=cd.get((r,c),1.0)
                cell.fill=FILL_LOW_CONF if conf_val<0.55 else FILL_DATA
                if c in NUMERIC_COLS:
                    cell.alignment=Alignment(horizontal='right',vertical='center',wrap_text=True)
                else:
                    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                if c==1: cell.font=Font(bold=True)
        ws_dst=out_wb.create_sheet(title=side)
        for row in ws_src.iter_rows():
            for sc in row:
                dc=ws_dst.cell(row=sc.row,column=sc.column,value=sc.value)
                if sc.has_style:
                    dc.font=copy.copy(sc.font); dc.border=copy.copy(sc.border)
                    dc.fill=copy.copy(sc.fill); dc.alignment=copy.copy(sc.alignment)
                    dc.number_format=sc.number_format; dc.protection=copy.copy(sc.protection)
        for k,cd2 in ws_src.column_dimensions.items(): ws_dst.column_dimensions[k].width=cd2.width
        for k,rd  in ws_src.row_dimensions.items():    ws_dst.row_dimensions[k].height=rd.height
        print(f'  Sheet {side} | {len(cells)} data cells written.')
    out_wb.save(out_path)
    if hasattr(out_path, 'seek'):
        out_path.seek(0)
    else:
        print(f'  Saved -> {out_path}')

def table_sheet_name(table_id, total_tables):
    if total_tables==2:
        return {1:'LeftTable',2:'RightTable'}.get(table_id,f'Table{table_id}')
    return f'Table{table_id}'

def table_display_name(table_id, total_tables):
    if total_tables==2:
        return {1:'Left',2:'Right'}.get(table_id,f'Table{table_id}')
    return f'Table{table_id}'

def build_rows_from_cells(cells, conf_dict):
    if not cells: return []
    max_r=max(r for r,c in cells)
    rows=[]
    for r in range(max_r+1):
        row={name:'' for name in COL_NAMES}
        has_value=False
        for c,name in COL_NAME_BY_INDEX.items():
            val=cells.get((r,c),'')
            if val=='':
                continue
            row[name]=val
            has_value=True
            conf=conf_dict.get((r,c))
            if conf is not None:
                row[f'{name}_conf']=round(float(conf),4)
        if has_value:
            rows.append(row)
    return rows

def rows_to_cells(rows):
    cells={}; conf_dict={}
    for r,row in enumerate(rows or []):
        for c,name in COL_NAME_BY_INDEX.items():
            val=str(row.get(name,'')).strip()
            if not val:
                continue
            cells[(r,c)]=val
            conf_key=f'{name}_conf'
            if conf_key in row and row[conf_key] not in ('', None):
                try:
                    conf_dict[(r,c)]=float(row[conf_key])
                except Exception:
                    pass
    return cells, conf_dict



def process_table_image(table_id, tbgr, total_tables):
    label=table_display_name(table_id,total_tables)
    print(f'[3] Table {table_id} ({label}) - {tbgr.shape[1]}x{tbgr.shape[0]}px')
    hlines, vlines = detect_grid(tbgr)
    nrows=len(hlines)-1; ncols=len(vlines)-1
    print(f'    Grid raw: {nrows}x{ncols}')
    vlines = enforce_min_columns(vlines, MIN_COLS_EXPECTED)
    ncols  = len(vlines)-1
    print(f'    Grid enforced: {nrows}x{ncols}')
    raw_results, nrows, ncols = extract_all_cells(tbgr, hlines, vlines)
    conf_dict = {(r,c):info['conf'] for (r,c),info in raw_results.items()}
    cleaned   = postprocess_grid(raw_results, nrows, ncols)
    skip      = detect_data_start(raw_results, nrows)
    cleaned, conf_dict = apply_scan_header_skip(cleaned, conf_dict, skip)
    skip=min(skip, max(0, len(hlines)-2))
    if skip>0:
        hlines=hlines[skip:]
    ncd  = max((c for r,c in cleaned), default=len(COL_NAMES)-1)+1
    cleaned, conf_dict, extra_skip = trim_leading_non_data_rows(cleaned, conf_dict, ncd)
    extra_skip=min(extra_skip, max(0, len(hlines)-2))
    if extra_skip>0:
        hlines=hlines[extra_skip:]
    ncd  = max((c for r,c in cleaned), default=len(COL_NAMES)-1)+1
    last = find_data_extent(cleaned, ncd, max_empty_streak=8)
    nb4  = len(cleaned)
    cleaned = {k:v for k,v in cleaned.items() if k[0]<=last+2}
    conf_dict = {k:v for k,v in conf_dict.items() if k in cleaned}
    print(f'    Extent 0-{last} | trimmed {nb4-len(cleaned)} garbage cells')
    cleaned, conf_dict, hlines, vlines = trim_data_region(
        cleaned, conf_dict, hlines, vlines, min_row_fill=0.05, min_col_fill=0.05
    )
    nrows=len(hlines)-1; ncols=len(vlines)-1
    print(f'    Final {nrows}x{ncols} | {len(cleaned)} cells')
    rows=build_rows_from_cells(cleaned, conf_dict)
    return {
        'table_id': table_id,
        'sheet_name': table_sheet_name(table_id, total_tables),
        'grid_size': f'{nrows}x{ncols}',
        'data_rows': len(rows),
        'rows': rows,
        'column_widths_px': [int(vlines[i+1]-vlines[i]) for i in range(ncols)],
        'row_heights_px': [int(hlines[i+1]-hlines[i]) for i in range(nrows)],
    }

def process_logbook_image(image_bytes, source_name='uploaded image'):
    init_ocr_engine()
    if not image_bytes:
        return {'error': 'No image data provided.'}
    arr=np.frombuffer(image_bytes,dtype=np.uint8)
    bgr=cv2.imdecode(arr,cv2.IMREAD_COLOR)
    if bgr is None:
        return {'error': 'Invalid image file.'}
    orig_h, orig_w = bgr.shape[:2]
    print('='*62)
    print(f'  Input : {source_name}')
    print('='*62)
    print(f'[1] Image: {orig_w}x{orig_h}px')
    bgr = rectify_document(bgr)
    print(f'    Rectified: {bgr.shape[1]}x{bgr.shape[0]}px')
    layout, tables = auto_split_tables(bgr)
    total_tables=len(tables)
    print(f'[2] Layout: {layout} | Tables: {total_tables}')
    result_tables={}
    for table_id, tbgr in tables:
        table_info=process_table_image(table_id, tbgr, total_tables)
        sheet_name=table_info.pop('sheet_name')
        result_tables[sheet_name]=table_info
    return {
        'layout': layout,
        'total_tables': total_tables,
        'image_size': f'{orig_w}x{orig_h}',
        'tables': result_tables,
    }

def export_to_excel(payload):
    if not isinstance(payload, dict):
        raise ValueError('OCR export payload must be a dictionary.')
    excel_file=payload.get('excel_file')
    if excel_file:
        excel_path=Path(excel_file)
        if excel_path.exists():
            return excel_path.read_bytes()
    tables=payload.get('tables') or {}
    if not tables:
        raise ValueError('No table data available for Excel export.')
    cells_by_table={}; conf_by_table={}; col_widths_by_table={}; row_hts_by_table={}
    for idx, (_, table_info) in enumerate(tables.items(), start=1):
        rows=table_info.get('rows') or []
        cells, conf_dict = rows_to_cells(rows)
        cells_by_table[idx]=cells
        conf_by_table[idx]=conf_dict
        if table_info.get('column_widths_px'):
            col_widths_by_table[idx]=list(table_info['column_widths_px'])
        if table_info.get('row_heights_px'):
            row_hts_by_table[idx]=list(table_info['row_heights_px'])
    buffer=BytesIO()
    write_filled_template(
        cells_by_table, conf_by_table, col_widths_by_table, row_hts_by_table, TEMPLATE_PATH, buffer
    )
    return buffer.getvalue()

def main():
    init_ocr_engine()
    images = (sorted(INPUT_FOLDER.glob('*.jpeg'))
              + sorted(INPUT_FOLDER.glob('*.jpg'))
              + sorted(INPUT_FOLDER.glob('*.png')))
    if not images:
        print(f'No images found in {INPUT_FOLDER}'); return
    for img_path in images:
        out_excel = OUTPUT_FOLDER / f'{img_path.stem}_filled.xlsx'
        result = process_logbook_image(img_path.read_bytes(), source_name=str(img_path))
        if 'error' in result:
            print(f"  ERROR: {result['error']}")
            continue
        print('[4] Writing Excel...')
        out_excel.write_bytes(export_to_excel(result))
        print(f'  Saved -> {out_excel}')

if __name__ == '__main__':
    main()
