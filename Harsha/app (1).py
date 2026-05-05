import os, time, re, warnings, copy, shutil, uuid
warnings.filterwarnings("ignore")

import cv2
import numpy as np
import pandas as pd
from pathlib import Path
from scipy.signal import find_peaks

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from paddleocr import PaddleOCR

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

app = FastAPI(title="OCR Pipeline API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── PATHS ──────────────────────────────────────────────────────────
UPLOAD_FOLDER   = "data/uploads"
OUTPUT_FOLDER   = "data/output"
TEMPLATE_PATH   = "data/template/template.xlsx"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ── OCR & CELL PARAMS ─────────────────────────────────────────────
OCR_MIN_CONF       = 0.20
CELL_TARGET_H_PX   = 120
CELL_MAX_SCALE     = 8.0
CELL_BORDER_STRIP  = 3
CELL_PAD_PX        = 8

NARROW_COLS        = {9, 10, 11, 12, 13}
NARROW_TARGET_H_PX = 160
NARROW_MAX_SCALE   = 12.0

BLEED_CONF_THRESHOLD = 0.32
BLEED_PATTERNS = [
    r"^[^a-zA-Z0-9]+$",
    r"^[.\-_,;:!?]{1,3}$",
    r"^.{1,2}$",
    r"[\\|/]{2,}",
]

NUMERIC_COLS      = {0, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}
TEXT_COLS         = {1, 5}
MIN_COLS_EXPECTED = 17
EXPECTED_ROWS_AFTER_SKIP = 65
TEMPLATE_HEADER_ROWS     = 6
TEMPLATE_TOTAL_LABEL     = "TOTAL"
DASH_ONLY = re.compile(r"^[\s.\-_,]+$")

_thin  = Side(style="thin",   color="000000")
_thick = Side(style="medium", color="000000")
BORDER_THIN  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
FILL_DATA    = PatternFill("solid", fgColor="FFFDE7")
FILL_TOTAL   = PatternFill("solid", fgColor="BBDEFB")
FILL_LOWCONF = PatternFill("solid", fgColor="FFF2CC")

def px_to_col_width(px): return max(1.0, (px - 5) / 7.0)
def px_to_row_ht(px):    return max(5.0, px * 0.75)

# ══════════════════════════════════════════════════════════════════
# FUZZY MATCHING
# ══════════════════════════════════════════════════════════════════

def levenshtein(s1, s2):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m, n = len(s1), len(s2)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[0], i
        for j in range(1, n + 1):
            prev, dp[j] = dp[j], (prev if s1[i-1] == s2[j-1]
                                   else 1 + min(prev, dp[j], dp[j-1]))
    return 1.0 - dp[n] / max(m, n)

OCR_CONFUSION = {
    ("b","s"):0.1, ("s","b"):0.1,
    ("0","o"):0.1, ("o","0"):0.1,
    ("1","l"):0.2, ("l","1"):0.2,
    ("1","i"):0.2, ("i","1"):0.2,
    ("u","n"):0.2, ("n","u"):0.2,
    ("c","e"):0.2, ("e","c"):0.2,
    ("a","u"):0.3, ("u","a"):0.3,
    ("v","y"):0.2, ("y","v"):0.2,
    ("m","n"):0.2, ("n","m"):0.2,
    ("d","a"):0.3, ("a","d"):0.3,
    ("f","t"):0.3, ("t","f"):0.3,
    ("p","r"):0.2, ("r","p"):0.2,
    ("g","q"):0.1, ("q","g"):0.1,
    ("6","b"):0.1, ("b","6"):0.1,
    ("4","a"):0.3, ("a","4"):0.3,
}

def weighted_edit_distance(s1, s2):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m, n = len(s1), len(s2)
    dp = [[0.0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1): dp[i][0] = float(i)
    for j in range(n + 1): dp[0][j] = float(j)
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = 0.0 if s1[i-1] == s2[j-1] else OCR_CONFUSION.get((s1[i-1], s2[j-1]), 1.0)
            dp[i][j] = min(dp[i-1][j] + 1.0, dp[i][j-1] + 1.0, dp[i-1][j-1] + cost)
    if m > 1 and n > 1 and s1[m-1] == s2[n-2] and s1[m-2] == s2[n-1]:
        dp[m][n] = min(dp[m][n], dp[m-2][n-2] + 0.5)
    return max(0.0, 1.0 - dp[m][n] / max(m, n))

def jaro_winkler(s1, s2, p=0.1, max_prefix=4):
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if not s1 or not s2: return 0.0
    m1, m2 = len(s1), len(s2)
    match_dist = max(m1, m2) // 2 - 1
    if match_dist < 0: match_dist = 0
    s1_matches = [False] * m1; s2_matches = [False] * m2
    matches = transpositions = 0
    for i in range(m1):
        lo = max(0, i - match_dist); hi = min(i + match_dist + 1, m2)
        for j in range(lo, hi):
            if s2_matches[j] or s1[i] != s2[j]: continue
            s1_matches[i] = s2_matches[j] = True; matches += 1; break
    if matches == 0: return 0.0
    k = 0
    for i in range(m1):
        if not s1_matches[i]: continue
        while not s2_matches[k]: k += 1
        if s1[i] != s2[k]: transpositions += 1
        k += 1
    jaro = (matches/m1 + matches/m2 + (matches - transpositions/2) / matches) / 3
    prefix = 0
    for i in range(min(max_prefix, m1, m2)):
        if s1[i] == s2[i]: prefix += 1
        else: break
    return jaro + prefix * p * (1 - jaro)

def qgram_similarity(s1, s2, q=2):
    from collections import Counter
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    if len(s1) < q or len(s2) < q: return float(s1 == s2)
    def grams(s): return [s[i:i+q] for i in range(len(s) - q + 1)]
    g1, g2 = grams(s1), grams(s2)
    c1, c2 = Counter(g1), Counter(g2)
    inter = sum((c1 & c2).values())
    return 2.0 * inter / (len(g1) + len(g2)) if (len(g1) + len(g2)) > 0 else 0.0

def permterm_similarity(s1, s2):
    from collections import Counter
    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2: return 1.0
    c1, c2 = Counter(s1), Counter(s2)
    inter = sum((c1 & c2).values())
    union = sum((c1 | c2).values())
    return inter / union if union > 0 else 0.0

FUZZY_WEIGHTS = {
    "levenshtein":   0.25,
    "weighted_edit": 0.35,
    "jaro_winkler":  0.20,
    "qgram":         0.12,
    "permterm":      0.08,
}

def fuzzy_score(ocr_text, candidate):
    return (FUZZY_WEIGHTS["levenshtein"]   * levenshtein(ocr_text, candidate) +
            FUZZY_WEIGHTS["weighted_edit"] * weighted_edit_distance(ocr_text, candidate) +
            FUZZY_WEIGHTS["jaro_winkler"]  * jaro_winkler(ocr_text, candidate) +
            FUZZY_WEIGHTS["qgram"]         * qgram_similarity(ocr_text, candidate) +
            FUZZY_WEIGHTS["permterm"]      * permterm_similarity(ocr_text, candidate))

def fuzzy_best_match(ocr_text, vocabulary, min_score=0.40):
    if not ocr_text: return None, 0.0
    t = ocr_text.strip().lower()
    scores = [(cand, fuzzy_score(t, cand.lower())) for cand in vocabulary]
    scores.sort(key=lambda x: -x[1])
    best_cand, best_sc = scores[0]
    return (best_cand, best_sc) if best_sc >= min_score else (None, 0.0)

MONTH_VOCAB = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","PR","PRN"]
MONTH_VOCAB_EXTENDED = MONTH_VOCAB + [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
MONTH_CANONICAL = {
    "January":"Jan","February":"Feb","March":"Mar","April":"Apr",
    "May":"May","June":"Jun","July":"Jul","August":"Aug",
    "September":"Sep","October":"Oct","November":"Nov","December":"Dec",
    "Jan":"Jan","Feb":"Feb","Mar":"Mar","Apr":"Apr",
    "Jun":"Jun","Jul":"Jul","Aug":"Aug","Sep":"Sep",
    "Oct":"Oct","Nov":"Nov","Dec":"Dec","PR":"PR","PRN":"PRN",
}

FERT_VOCAB = ["U-709","U-330","U-428","T-750","B/A","B/N","DAP","MOP","PH-B","PH-4.99A","PH-5.24A"]
FERT_NUM_PAT   = re.compile(r"\b([Uu][_\-]?|[Tt][_\-]?)(\d{2,4})\b")
FERT_KNOWN_NUMS = {"709":"U-709","330":"U-330","428":"U-428","750":"T-750"}

def fuzzy_match_month(ocr_text):
    t = str(ocr_text).strip()
    if not t: return t
    tl = t.lower()
    for m in MONTH_VOCAB_EXTENDED:
        if tl == m.lower(): return MONTH_CANONICAL.get(m, m[:3].title())
    if len(tl) >= 3:
        for m in MONTH_VOCAB_EXTENDED:
            if m.lower().startswith(tl[:3]) or tl[:3] == m.lower()[:3]:
                return MONTH_CANONICAL.get(m, m[:3].title())
    best, sc = fuzzy_best_match(tl, MONTH_VOCAB_EXTENDED, min_score=0.42)
    if best:
        return MONTH_CANONICAL.get(best, best[:3].title())
    return t[:3].title() if t[:3].isalpha() else t

def fuzzy_match_fert(ocr_text):
    t = str(ocr_text).strip()
    if not t: return t
    m = FERT_NUM_PAT.search(t)
    if m:
        num = m.group(2); prefix = m.group(1).upper().rstrip("-_")
        if num in FERT_KNOWN_NUMS: return FERT_KNOWN_NUMS[num]
        return f"{prefix}-{num}"
    digits = re.findall(r"\d{3,4}", t)
    for d in digits:
        if d in FERT_KNOWN_NUMS: return FERT_KNOWN_NUMS[d]
    best, sc = fuzzy_best_match(t, FERT_VOCAB, min_score=0.38)
    if best: return best
    tl = t.lower()
    if re.search(r"b[/\\]?a", tl): return "B/A"
    if re.search(r"b[/\\]?n", tl): return "B/N"
    if "dap" in tl: return "DAP"
    if "mop" in tl: return "MOP"
    if re.search(r"ph[-\s]?b", tl): return "PH-B"
    return " ".join(t.split())

# ══════════════════════════════════════════════════════════════════
# OCR ENGINE
# ══════════════════════════════════════════════════════════════════

def make_ocr_engine():
    for kwargs in [
        dict(lang="en", det_model_name="PP-OCRv5_server_det",
             rec_model_name="PP-OCRv5_server_rec", use_textline_orientation=False),
        dict(lang="en", ocr_version="PP-OCRv5", use_textline_orientation=False),
        dict(lang="en", use_textline_orientation=False),
        dict(lang="en"),
    ]:
        try:
            eng = PaddleOCR(**kwargs)
            return eng
        except Exception:
            continue
    raise RuntimeError("PaddleOCR init failed")

OCR_ENGINE = make_ocr_engine()

def _parse_paddle_output(out):
    items = []
    if not out: return items
    if isinstance(out, list) and len(out) > 0 and isinstance(out[0], dict):
        d = out[0]
        texts  = d.get("rec_texts") or d.get("texts") or []
        scores = d.get("rec_scores") or d.get("scores") or ([1.0] * len(texts))
        boxes  = (d.get("dt_polys") or d.get("rec_polys") or
                  d.get("dt_boxes") or d.get("boxes") or [])
        if isinstance(boxes, np.ndarray): boxes = boxes.tolist()
        if isinstance(texts, str): texts = [texts]
        for b, t, s in zip(boxes, texts, scores):
            items.append((b, t, float(s)))
        return items
    lines = out[0] if (isinstance(out[0], list) and not isinstance(out[0][0], list)) else out
    for line in lines:
        try: items.append((line[0], line[1][0], float(line[1][1])))
        except Exception: pass
    return items

def run_ocr(img_bgr):
    if img_bgr is None or img_bgr.size == 0: return []
    if img_bgr.ndim == 2: img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)
    try:
        if hasattr(OCR_ENGINE, "predict"):
            try:
                out = OCR_ENGINE.predict(img_bgr,
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_textline_orientation=False)
            except TypeError:
                out = OCR_ENGINE.predict(img_bgr)
        else:
            out = OCR_ENGINE.ocr(img_bgr)
        return _parse_paddle_output(out)
    except Exception:
        return []

# ══════════════════════════════════════════════════════════════════
# IMAGE & GRID
# ══════════════════════════════════════════════════════════════════

def _four_point_warp(bgr, pts):
    pts = pts.reshape(4, 2); s = pts.sum(1); d = np.diff(pts, axis=1)
    rect = np.array([pts[s.argmin()], pts[d.argmin()],
                     pts[s.argmax()], pts[d.argmax()]], dtype=np.float32)
    tl, tr, br, bl = rect
    W = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
    H = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
    dst = np.array([[0,0],[W-1,0],[W-1,H-1],[0,H-1]], dtype=np.float32)
    return cv2.warpPerspective(bgr, cv2.getPerspectiveTransform(rect, dst), (W, H))

def rectify_document(bgr):
    gray  = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(cv2.GaussianBlur(gray, (5, 5), 0), 50, 150)
    edges = cv2.dilate(edges, np.ones((5, 5), np.uint8))
    cnts, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts: return bgr
    for c in sorted(cnts, key=cv2.contourArea, reverse=True)[:5]:
        ap = cv2.approxPolyDP(c, 0.02 * cv2.arcLength(c, True), True)
        if len(ap) == 4:
            return _four_point_warp(bgr, ap)
    return bgr

def split_left_right(bgr, pad=3):
    h, w = bgr.shape[:2]
    _, bi = cv2.threshold(cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY), 200, 255, cv2.THRESH_BINARY_INV)
    vk = max(30, h // 18)
    vm = cv2.morphologyEx(bi, cv2.MORPH_OPEN,
                          cv2.getStructuringElement(cv2.MORPH_RECT, (1, vk)))
    proj = np.convolve(
        np.sum(vm[int(h * 0.2):, :] > 0, axis=0).astype(float),
        np.ones(max(15, w // 200)) / max(15, w // 200), mode="same")
    if proj.max() == 0: return [(1, bgr)]
    peaks, _ = find_peaks(proj, height=proj.max() * 0.25, distance=max(12, w // 100))
    if len(peaks) == 0: return [(1, bgr)]
    mid = peaks[np.argmin(np.abs(peaks - w * 0.5))]
    tables = [(i, bgr[:, max(0, x1 - pad):min(w, x2 + pad)].copy())
              for i, (x1, x2) in enumerate([(0, mid), (mid, w)], 1)
              if x2 - x1 > w * 0.1]
    return tables if len(tables) == 2 else [(1, bgr)]

def _cluster(pos, eps=8):
    pos = np.sort(np.unique(pos.astype(int)))
    if len(pos) == 0: return pos
    groups, cur = [[pos[0]]], pos[0]
    for p in pos[1:]:
        if p - cur <= eps: groups[-1].append(p)
        else: groups.append([p])
        cur = p
    return np.array([int(np.mean(g)) for g in groups])

def detect_grid(bgr):
    h, w = bgr.shape[:2]
    g  = cv2.createCLAHE(2.0, (8, 8)).apply(cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY))
    th = cv2.adaptiveThreshold(g, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                               cv2.THRESH_BINARY_INV, 31, 15)
    th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8), iterations=2)
    hk = max(20, w // 30); vk = max(20, h // 30)
    hl = cv2.dilate(cv2.morphologyEx(th, cv2.MORPH_OPEN,
                    cv2.getStructuringElement(cv2.MORPH_RECT, (hk, 1))),
                    np.ones((1, 3), np.uint8))
    vl = cv2.dilate(cv2.morphologyEx(th, cv2.MORPH_OPEN,
                    cv2.getStructuringElement(cv2.MORPH_RECT, (1, vk))),
                    np.ones((3, 1), np.uint8))
    ys = [y + hh // 2 for c in cv2.findContours(hl, cv2.RETR_EXTERNAL,
          cv2.CHAIN_APPROX_SIMPLE)[0]
          for x, y, ww, hh in [cv2.boundingRect(c)] if ww >= w * 0.40]
    xs = [x + ww // 2 for c in cv2.findContours(vl, cv2.RETR_EXTERNAL,
          cv2.CHAIN_APPROX_SIMPLE)[0]
          for x, y, ww, hh in [cv2.boundingRect(c)] if hh >= h * 0.40]
    ys = _cluster(np.array(ys)); xs = _cluster(np.array(xs))
    def _border(lines, lo, hi, tol=14):
        l = list(lines)
        if not l or abs(l[0] - lo) > tol: l = [lo] + l
        if abs(l[-1] - hi) > tol: l += [hi]
        return np.array(sorted(l))
    return _border(ys, 0, h - 1), _border(xs, 0, w - 1)

def enforce_min_columns(vlines, min_cols=MIN_COLS_EXPECTED):
    vl = list(vlines)
    while len(vl) - 1 < min_cols:
        widths = [(vl[i+1] - vl[i], i) for i in range(len(vl) - 1)]
        _, idx = max(widths)
        mid = (vl[idx] + vl[idx + 1]) // 2
        vl.insert(idx + 1, mid)
    return np.array(vl, dtype=float)

def split_three_tables(bgr, pad=3):
    h, w = bgr.shape[:2]
    _, bi = cv2.threshold(cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY), 200, 255, cv2.THRESH_BINARY_INV)
    vk = max(int(h / 20), 30)
    vert_clean = cv2.morphologyEx(bi, cv2.MORPH_OPEN,
                                  cv2.getStructuringElement(cv2.MORPH_RECT, (1, vk)))
    proj = np.sum(vert_clean, axis=0).astype(float)
    if proj.max() == 0: return [(1, bgr)]
    peaks, _ = find_peaks(proj, height=proj.max() * 0.10, distance=5)
    zones = {
        "A": (int(w * 0.25), int(w * 0.40)),
        "B": (int(w * 0.45), int(w * 0.55)),
        "C": (int(w * 0.60), int(w * 0.75)),
    }
    noise_floor = w * 0.015
    best = {}
    for zn, (zx0, zx1) in zones.items():
        zp = [p for p in peaks if zx0 <= p <= zx1]
        if len(zp) < 2:
            best[zn] = {"size": 0, "cut_x": 0}; continue
        gaps = np.diff(zp); idx = int(np.argmax(gaps))
        best[zn] = {"size": int(gaps[idx]), "cut_x": int(zp[idx + 1])}
    cuts = []
    if best["A"]["size"] > noise_floor and best["A"]["size"] >= best["B"]["size"]:
        cuts.append(best["A"]["cut_x"])
        cuts.append(best["C"]["cut_x"] if best["C"]["size"] > noise_floor else int(w * 0.66))
    else:
        cx = best["B"]["cut_x"] if best["B"]["size"] > noise_floor else int(w * 0.50)
        cuts.append(cx)
    boundaries = [0] + sorted(cuts) + [w]
    tables = []
    for i in range(len(boundaries) - 1):
        x1, x2 = boundaries[i], boundaries[i + 1]
        if (x2 - x1) > w * 0.10:
            tables.append((i + 1, bgr[:, max(0, x1 - pad):min(w, x2 + pad)].copy()))
    return tables if tables else [(1, bgr)]

def auto_split_tables(bgr, pad=3):
    h, w = bgr.shape[:2]
    _, bi = cv2.threshold(cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY), 200, 255, cv2.THRESH_BINARY_INV)
    vk = max(int(h / 20), 30)
    vert_clean = cv2.morphologyEx(bi, cv2.MORPH_OPEN,
                                  cv2.getStructuringElement(cv2.MORPH_RECT, (1, vk)))
    proj = np.sum(vert_clean, axis=0).astype(float)
    if proj.max() == 0: return "2-table", [(1, bgr)]
    peaks, _ = find_peaks(proj, height=proj.max() * 0.10, distance=5)
    noise_floor = w * 0.015
    def _zone_gap(zx0, zx1):
        zp = sorted(p for p in peaks if zx0 <= p <= zx1)
        if len(zp) < 2: return 0
        return int(np.max(np.diff(zp)))
    gap_A = _zone_gap(int(w * 0.25), int(w * 0.40))
    gap_B = _zone_gap(int(w * 0.45), int(w * 0.55))
    if gap_A > noise_floor and gap_A >= gap_B:
        return "3-table", split_three_tables(bgr, pad=pad)
    return "2-table", split_left_right(bgr, pad=pad)

# ══════════════════════════════════════════════════════════════════
# CELL-BY-CELL OCR
# ══════════════════════════════════════════════════════════════════

def is_bleed_through(text, conf, col_idx):
    t = str(text).strip()
    if col_idx in NUMERIC_COLS and DASH_ONLY.match(t): return True
    if conf < BLEED_CONF_THRESHOLD:
        for pattern in BLEED_PATTERNS:
            if re.search(pattern, t): return True
    if col_idx == 1:
        if len(t) > 8 or (len(t) <= 2 and conf < 0.4): return True
    if col_idx in NUMERIC_COLS:
        if not any(c.isdigit() for c in t) and conf < 0.4: return True
    return False

def _scale_up(gray, target_h=CELL_TARGET_H_PX, max_scale=CELL_MAX_SCALE):
    h = gray.shape[0]
    if h < 4: return gray
    scale = min(max_scale, max(1.0, target_h / h))
    if scale <= 1.01: return gray
    return cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

def _scale_up_narrow(gray):
    h = gray.shape[0]
    if h < 4: return gray
    scale = min(NARROW_MAX_SCALE, max(1.0, NARROW_TARGET_H_PX / h))
    if scale <= 1.01: return gray
    return cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_LANCZOS4)

def _prep_numeric(gray):
    g = cv2.createCLAHE(2.0, (4, 4)).apply(gray)
    _, b = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if np.mean(b) < 127: b = cv2.bitwise_not(b)
    b = cv2.fastNlMeansDenoising(b, h=10)
    b = cv2.morphologyEx(b, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))
    return cv2.filter2D(b, -1, np.array([[-1,-1,-1],[-1,9,-1],[-1,-1,-1]]))

def _prep_text(gray):
    g = cv2.createCLAHE(3.0, (4, 4)).apply(gray)
    b = cv2.adaptiveThreshold(g, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                              cv2.THRESH_BINARY, 21, 10)
    if np.mean(b) < 127: b = cv2.bitwise_not(b)
    return cv2.morphologyEx(b, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))

def _prep_clahe_only(gray):
    g = cv2.createCLAHE(2.5, (4, 4)).apply(gray)
    return cv2.filter2D(g, -1, np.array([[0,-1,0],[-1,5,-1],[0,-1,0]]))

def _prep_narrow_col(gray):
    g = cv2.createCLAHE(3.0, (4, 4)).apply(gray)
    g = cv2.bilateralFilter(g, 5, 75, 75)
    _, b = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if np.mean(b) < 127: b = cv2.bitwise_not(b)
    b = cv2.morphologyEx(b, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))
    return cv2.filter2D(b, -1, np.array([[-1,-1,-1],[-1,9,-1],[-1,-1,-1]]))

def crop_cell(bgr, hlines, vlines, row, col):
    H, W = bgr.shape[:2]; s = CELL_BORDER_STRIP
    y1 = min(H - 1, int(hlines[row]) + s);   y2 = max(0, int(hlines[row + 1]) - s)
    x1 = min(W - 1, int(vlines[col]) + s);   x2 = max(0, int(vlines[col + 1]) - s)
    if y2 - y1 < 4 or x2 - x1 < 4: return None
    crop = bgr[y1:y2, x1:x2].copy()
    p = CELL_PAD_PX
    return cv2.copyMakeBorder(crop, p, p, p, p, cv2.BORDER_CONSTANT, value=(255, 255, 255))

def ocr_single_cell(bgr_crop, col_idx):
    if bgr_crop is None or bgr_crop.size == 0: return "", 0.0
    gray = cv2.cvtColor(bgr_crop, cv2.COLOR_BGR2GRAY)
    if col_idx in NARROW_COLS:
        gray = _scale_up_narrow(gray)
        v1, v2, v3 = _prep_narrow_col(gray), _prep_numeric(gray), _prep_clahe_only(gray)
    elif col_idx in NUMERIC_COLS:
        gray = _scale_up(gray)
        v1 = _prep_numeric(gray); v2 = _prep_clahe_only(gray)
        cg = cv2.createCLAHE(2.0, (4, 4)).apply(gray)
        _, v3 = cv2.threshold(cg, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        if np.mean(v3) < 127: v3 = cv2.bitwise_not(v3)
    else:
        gray = _scale_up(gray)
        v1 = _prep_text(gray); v2 = _prep_clahe_only(gray)
        cg = cv2.createCLAHE(2.0, (4, 4)).apply(gray)
        _, v3 = cv2.threshold(cg, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        if np.mean(v3) < 127: v3 = cv2.bitwise_not(v3)
    best_text, best_conf = "", 0.0
    for v in (v1, v2, v3):
        img   = cv2.cvtColor(v, cv2.COLOR_GRAY2BGR) if v.ndim == 2 else v
        items = run_ocr(img)
        good  = [(t, s) for _, t, s in items if s >= OCR_MIN_CONF]
        if not good: continue
        text = " ".join(t for t, _ in good).strip()
        conf = float(np.mean([s for _, s in good]))
        if conf > best_conf: best_conf, best_text = conf, text
    if best_text and is_bleed_through(best_text, best_conf, col_idx):
        return "", 0.0
    return best_text, best_conf

def extract_all_cells(bgr, hlines, vlines):
    nrows = len(hlines) - 1; ncols = len(vlines) - 1
    results = {}; t0 = time.time()
    for r in range(nrows):
        for c in range(ncols):
            crop = crop_cell(bgr, hlines, vlines, r, c)
            text, conf = ocr_single_cell(crop, c)
            if text: results[(r, c)] = {"text": text, "conf": conf}
    return results, nrows, ncols

# ══════════════════════════════════════════════════════════════════
# POST-PROCESSING
# ══════════════════════════════════════════════════════════════════

def clean_numeric(text):
    text = str(text).strip()
    if not text or DASH_ONLY.match(text): return ""
    spaced = re.sub(r"(?<=\d)\s+(?=\d)", "", text)
    nums = re.findall(r"\d+(?:[.,]\d+)?", spaced)
    if nums: return nums[0] if len(nums) == 1 else " ".join(nums)
    nums2 = re.findall(r"\d+(?:[.,]\d+)?", text)
    if nums2: return nums2[0] if len(nums2) == 1 else " ".join(nums2)
    return "".join(c for c in text if c.isdigit() or c in ".,/-+").strip()

def clean_cell(text, col_idx):
    text = str(text).strip()
    if not text: return ""
    if col_idx in NUMERIC_COLS and DASH_ONLY.match(text): return ""
    if col_idx in NUMERIC_COLS: return clean_numeric(text)
    if col_idx == 1: return fuzzy_match_month(text)
    if col_idx == 5: return fuzzy_match_fert(text)
    return " ".join(text.split())

def post_process_grid(raw_results, nrows, ncols):
    cleaned = {}
    for (r, c), info in raw_results.items():
        val = clean_cell(info["text"], c)
        if val: cleaned[(r, c)] = val
    return cleaned

def detect_data_start(raw_results, nrows, max_scan=15, fallback=5):
    skip_grid = max(0, nrows - EXPECTED_ROWS_AFTER_SKIP)
    NO_PAT    = re.compile(r"^\d{1,2}$")
    MONTH_PAT = re.compile(
        r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec"
        r"|Bep|Nnv|Auu|Jau|Fcd|Col|Mav|Tun|Sct|Aot|Nec|Fob|Map|Seo"
        r"|PR|PRN)", re.IGNORECASE)
    skip_ocr = None
    for r in range(min(nrows, max_scan)):
        t0 = str(raw_results.get((r, 0), {}).get("text", "")).strip()
        t1 = str(raw_results.get((r, 1), {}).get("text", "")).strip()
        has_no    = bool(NO_PAT.match(t0))
        has_month = bool(MONTH_PAT.match(t1))
        n_filled  = sum(1 for c in range(17)
                        if raw_results.get((r, c), {}).get("text", "").strip())
        triggered = ((has_no and has_month) or
                     (has_no and n_filled >= 3) or
                     (has_month and n_filled >= 4))
        if triggered:
            row_num  = int(t0) if (has_no and t0.isdigit()) else 1
            skip_ocr = max(0, r - (row_num - 1))
            break
    if skip_ocr is None: skip_ocr = fallback
    if skip_grid <= 0: skip = skip_ocr
    elif abs(skip_grid - skip_ocr) <= 1: skip = skip_grid
    else: skip = skip_ocr
    return skip

def apply_scan_header_skip(cells, conf_dict, skip):
    new_cells = {}; new_conf = {}
    for (r, c), val in cells.items():
        if r < skip: continue
        r2 = r - skip
        new_cells[(r2, c)] = val
        if (r, c) in conf_dict: new_conf[(r2, c)] = conf_dict[(r, c)]
    return new_cells, new_conf

def find_data_extent(cells, ncols, max_empty_streak=8):
    if not cells: return 0
    max_r = max(r for r, c in cells)
    last_valid = -1; streak = 0
    for r in range(max_r + 1):
        month_v = str(cells.get((r, 1), "")).strip()
        no_v    = str(cells.get((r, 0), "")).strip()
        has_no    = bool(re.fullmatch(r"\d{1,3}", no_v))
        has_month = bool(re.fullmatch(r"[A-Z][a-z]{1,4}|PR|PRN|TOTAL", month_v))
        is_total  = "TOTAL" in month_v.upper() or "TOTAL" in no_v.upper()
        n_cells   = sum(1 for c in range(ncols) if cells.get((r, c), ""))
        if (has_no or has_month or is_total) and n_cells >= 1:
            last_valid = r; streak = 0
        else:
            streak += 1
            if streak >= max_empty_streak and last_valid >= 0: break
    return max(last_valid, 0)

def trim_data_region(cells, hlines, vlines, min_row_fill=0.05, min_col_fill=0.05):
    nrows = len(hlines) - 1; ncols = len(vlines) - 1
    if nrows == 0 or ncols == 0: return cells, hlines, vlines
    row_cnt = [sum(1 for c in range(ncols) if cells.get((r, c), "")) for r in range(nrows)]
    col_cnt = [sum(1 for r in range(nrows) if cells.get((r, c), "")) for c in range(ncols)]
    r_thresh = max(1, int(ncols * min_row_fill))
    c_thresh = max(1, int(nrows * min_col_fill))
    act_r = [i for i, v in enumerate(row_cnt) if v >= r_thresh]
    act_c = [i for i, v in enumerate(col_cnt) if v >= c_thresh]
    if not act_r or not act_c: return cells, hlines, vlines
    r0, r1 = min(act_r), max(act_r)
    c0, c1 = min(act_c), max(act_c)
    new_cells = {(r - r0, c - c0): v for (r, c), v in cells.items()
                 if r0 <= r <= r1 and c0 <= c <= c1}
    return new_cells, hlines[r0:r1+2], vlines[c0:c1+2]

# ══════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ══════════════════════════════════════════════════════════════════

def write_filled_template(cells_by_table, conf_by_table,
                          col_widths_by_table, row_hts_by_table,
                          template_path, out_path):
    n_tables = len(cells_by_table)
    if n_tables <= 2:
        side_labels = {1: "Left_Table", 2: "Right_Table"}
    else:
        side_labels = {i: f"Table_{i}" for i in cells_by_table}
    out_wb = Workbook(); out_wb.remove(out_wb.active)
    for t_idx, cells in cells_by_table.items():
        side = side_labels.get(t_idx, f"Table_{t_idx}")
        try:
            tmpl_wb = load_workbook(template_path); ws_src = tmpl_wb.active
        except Exception:
            tmpl_wb = Workbook(); ws_src = tmpl_wb.active
        if t_idx in col_widths_by_table:
            for ci, w in enumerate(col_widths_by_table[t_idx]):
                ws_src.column_dimensions[get_column_letter(ci + 1)].width = px_to_col_width(w)
        if t_idx in row_hts_by_table:
            for ri, h in enumerate(row_hts_by_table[t_idx]):
                ws_src.row_dimensions[TEMPLATE_HEADER_ROWS + ri + 1].height = px_to_row_ht(h)
        cd = conf_by_table.get(t_idx, {})
        for (r, c), val in cells.items():
            excel_row = TEMPLATE_HEADER_ROWS + r + 1
            if excel_row <= TEMPLATE_HEADER_ROWS: continue
            cell = ws_src.cell(row=excel_row, column=c + 1, value=val)
            cell.border = BORDER_THIN
            is_total = str(val).upper().strip() == TEMPLATE_TOTAL_LABEL
            if is_total:
                cell.fill = FILL_TOTAL; cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                conf_val = cd.get((r, c), 1.0)
                cell.fill = FILL_LOWCONF if conf_val < 0.55 else FILL_DATA
                if c in NUMERIC_COLS:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if c == 1: cell.font = Font(bold=True)
        ws_dst = out_wb.create_sheet(title=side)
        for row in ws_src.iter_rows():
            for sc in row:
                dc = ws_dst.cell(row=sc.row, column=sc.column, value=sc.value)
                if sc.has_style:
                    dc.font = copy.copy(sc.font)
                    dc.border = copy.copy(sc.border)
                    dc.fill = copy.copy(sc.fill)
                    dc.alignment = copy.copy(sc.alignment)
                    dc.number_format = sc.number_format
                    dc.protection = copy.copy(sc.protection)
        for k, cd2 in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[k].width = cd2.width
        for k, rd in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[k].height = rd.height
    out_wb.save(out_path)

# ══════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════

def run_pipeline(img_path: str, out_excel: str) -> dict:
    bgr = cv2.imread(img_path)
    if bgr is None: raise FileNotFoundError(f"Image not found: {img_path}")
    bgr = rectify_document(bgr)
    layout, tables = auto_split_tables(bgr)
    all_cells: dict = {}; all_confs: dict = {}
    col_widths_px: dict = {}; row_hts_px: dict = {}
    for t_idx, t_bgr in tables:
        hlines, vlines = detect_grid(t_bgr)
        vlines = enforce_min_columns(vlines, MIN_COLS_EXPECTED)
        nrows = len(hlines) - 1; ncols = len(vlines) - 1
        raw_results, nrows, ncols = extract_all_cells(t_bgr, hlines, vlines)
        conf_dict = {(r, c): info["conf"] for (r, c), info in raw_results.items()}
        cleaned   = post_process_grid(raw_results, nrows, ncols)
        skip = detect_data_start(raw_results, nrows)
        cleaned, conf_dict = apply_scan_header_skip(cleaned, conf_dict, skip)
        nc_d   = max((c for r, c in cleaned), default=0) + 1
        last   = find_data_extent(cleaned, nc_d, max_empty_streak=8)
        cleaned = {k: v for k, v in cleaned.items() if k[0] <= last + 2}
        cleaned, hlines, vlines = trim_data_region(cleaned, hlines, vlines)
        nrows = len(hlines) - 1; ncols = len(vlines) - 1
        col_widths_px[t_idx] = [int(vlines[i+1] - vlines[i]) for i in range(ncols)]
        row_hts_px[t_idx]    = [int(hlines[i+1] - hlines[i]) for i in range(nrows)]
        all_cells[t_idx]  = cleaned
        all_confs[t_idx]  = conf_dict
    write_filled_template(all_cells, all_confs, col_widths_px, row_hts_px,
                          TEMPLATE_PATH, out_excel)
    total_cells = sum(len(v) for v in all_cells.values())
    return {
        "layout": layout,
        "tables": len(tables),
        "total_cells": total_cells,
        "output": out_excel,
    }

# ══════════════════════════════════════════════════════════════════
# FASTAPI ROUTES
# ══════════════════════════════════════════════════════════════════

@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/process")
async def process_image(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}:
        raise HTTPException(status_code=400,
                            detail="Unsupported file type. Use JPG, PNG, BMP or TIFF.")
    job_id   = uuid.uuid4().hex
    img_path = os.path.join(UPLOAD_FOLDER, f"{job_id}{ext}")
    out_path = os.path.join(OUTPUT_FOLDER, f"{job_id}_filled.xlsx")
    try:
        with open(img_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
        result = run_pipeline(img_path, out_path)
        return JSONResponse({
            "job_id":      job_id,
            "layout":      result["layout"],
            "tables":      result["tables"],
            "total_cells": result["total_cells"],
            "download_url": f"/download/{job_id}",
        })
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(img_path):
            os.remove(img_path)

@app.get("/download/{job_id}")
def download_result(job_id: str):
    out_path = os.path.join(OUTPUT_FOLDER, f"{job_id}_filled.xlsx")
    if not os.path.exists(out_path):
        raise HTTPException(status_code=404, detail="Result not found.")
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"ocr_result_{job_id}.xlsx",
    )

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=False)
